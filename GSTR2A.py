from openpyxl import Workbook , load_workbook
import pandas as pd 
import numpy as np

def reco_itr_2a(file1, file2, sheet_name):
    wb = Workbook()
    ws = wb.active
    # ws['A1'] = 'GSTIN/UIN'
    # ws['B1'] = 'Particulars'
    # ws['C1'] = 'Date'
    wb.save('CLIENT.xlsx')
    client = pd.read_excel(file1)
    govern = pd.read_excel(file2, sheet_name='B2B')
    ORIGINAL_GSTR2A = pd.read_excel(file2,sheet_name='B2B')

    row = {'GSTIN/UIN':'NOT AVAILABLE' , 'Particulars':'NOT AVAILABLE' , 'Date':'NOT AVAILABLE'}
    # resizing client side data

    resizing = 0
    if govern.shape[0] != client.shape[0]:
        resizing = govern.shape[0]-client.shape[0]
        print(resizing)



    for i in range(resizing):
        client = pd.concat([client,pd.DataFrame([row])])


    client['Particulars'] = client['Particulars'].str.lower()
    govern['Trade/Legal name'] = govern['Trade/Legal name'].str.lower()



    i = 0
    j = 0

    while i < len(govern) and j < len(client):
        print(client.iloc[j]['Particulars'])
        str_arr = govern.iloc[i]['Trade/Legal name'].split()
        if govern.iloc[i]['GSTIN of supplier'] == client.iloc[j]['GSTIN/UIN'] and  str_arr[0] in client.iloc[j]['Particulars'] and govern.iloc[i]['Invoice Date'] == client.iloc[j]['Date'] and govern.iloc[i]['Invoice Value(₹)'] == client.iloc[j]['Gross Total']:
            ws[f'A{i+1}'] = client.iloc[j]['GSTIN/UIN']
            ws[f'B{i+1}'] = client.iloc[j]['Particulars']
            ws[f'C{i+1}'] = client.iloc[j]['Date']
            ws[f'D{i+1}'] = client.iloc[j]['Supplier Invoice No.']
            ws[f'E{i+1}'] = client.iloc[j]['Gross Total']
            ws[f'F{i+1}'] = client.iloc[j]['C-GST 9%']
            ws[f'G{i+1}'] = client.iloc[j]['S-GST 9%']
            ws[f'H{i+1}'] = client.iloc[j]['I-GST 18%']
            # print('found..',govern.iloc[i]['GSTIN of supplier'], govern.iloc[i]['Trade/Legal name'] , client.iloc[index]['GSTIN/UIN'] ,  client.iloc[index]['Particulars'] )
            wb.save('CLIENT.xlsx')
        else:
            for index in range(len(client)):
                print(i , index)
                print(client.iloc[index]['Particulars'])
                print('current','searching...',govern.iloc[i]['GSTIN of supplier'], govern.iloc[i]['Trade/Legal name'] , client.iloc[index]['GSTIN/UIN'] ,  client.iloc[index]['Particulars'] )
                GSTIN_VALIDIATION = govern.iloc[i]['GSTIN of supplier'] == client.iloc[index]['GSTIN/UIN']
                SUPPLIER_VALIDIATION = str_arr[0] or str_arr[1] in client.iloc[index]['Particulars']
                DATE_VALIDIATION = govern.iloc[i]['Invoice Date'] == client.iloc[index]['Date']
                if (GSTIN_VALIDIATION and SUPPLIER_VALIDIATION  and DATE_VALIDIATION) or ( str_arr[0] in client.iloc[index]['Particulars'] and DATE_VALIDIATION or govern.iloc[i]['Invoice Value(₹)'] == client.iloc[index]['Gross Total'] ):
                    new_row = {'GSTIN/UIN':client.iloc[index]['GSTIN/UIN'], 'Particulars':client.iloc[index]['Particulars']}
                    print('searching...',govern.iloc[i]['GSTIN of supplier'], govern.iloc[i]['Trade/Legal name'],govern.iloc[i]['Invoice Date'] , client.iloc[index]['GSTIN/UIN'] ,  client.iloc[index]['Particulars'] ,  client.iloc[index]['Date'] )
                    ws[f'A{i+1}'] = client.iloc[index]['GSTIN/UIN']
                    ws[f'B{i+1}'] = client.iloc[index]['Particulars']
                    ws[f'C{i+1}'] = client.iloc[index]['Date']
                    ws[f'D{i+1}'] = client.iloc[index]['Supplier Invoice No.']
                    ws[f'E{i+1}'] = client.iloc[index]['Gross Total']
                    ws[f'F{i+1}'] = client.iloc[index]['C-GST 9%']
                    ws[f'G{i+1}'] = client.iloc[index]['S-GST 9%']
                    ws[f'H{i+1}'] = client.iloc[index]['I-GST 18%']
                    wb.save('CLIENT.xlsx')
                    break
                else:
                    new_row = {'GSTIN/UIN':client.iloc[index]['GSTIN/UIN'], 'Particulars':client.iloc[index]['Particulars']}
        j+=1
        i+=1

    ws.insert_rows(1) 
    ws['A1'] = 'GSTIN/UIN'
    ws['B1'] = 'Particulars'
    ws['C1'] = 'Date'
    ws['D1'] = 'Supplier Invoice No.'
    ws['E1'] = 'Gross Total'
    ws['F1'] = 'C-GST 9%'
    ws['G1'] = 'S-GST 9%'
    ws['H1'] = 'I-GST 18%'

    wb.save('CLIENT.xlsx')



    print(len(govern) , len(client))

    # Replace NaN values with 0 in GENERATED_PURCHASE_REGISTER



    GENERATED_PURCHASE_REGISTER = pd.read_excel('CLIENT.xlsx')
    GENERATED_PURCHASE_REGISTER['Particulars'] = GENERATED_PURCHASE_REGISTER['Particulars']
    ORIGINAL_GSTR2A['Trade/Legal name'] = ORIGINAL_GSTR2A['Trade/Legal name'].str.lower()
    GENERATED_PURCHASE_REGISTER['Particulars'] = GENERATED_PURCHASE_REGISTER['Particulars'].str.lower() 
    GENERATED_PURCHASE_REGISTER['Gross Total'].fillna(0 , inplace=True)


    remarks = []
    statement = ''


    wb = load_workbook(file2)
    ws = wb['B2B']




    for i in range(len(ORIGINAL_GSTR2A)):
        # print('PRINTING...' , ORIGINAL_GSTR2A.iloc[i]['Trade/Legal name'] , GENERATED_PURCHASE_REGISTER.iloc[i]['Particulars'])
        STRING_ARRAY = ORIGINAL_GSTR2A.iloc[i]['Trade/Legal name'].split()

        # if STRING_ARRAY[0] not in GENERATED_PURCHASE_REGISTER.iloc[i]['Particulars']:
        #     statement = statement +' '+ 'Supplier name ,'
        #     print(STRING_ARRAY[0] , GENERATED_PURCHASE_REGISTER.iloc[i]['Particulars'])
        if ORIGINAL_GSTR2A.iloc[i]['GSTIN of supplier'] != GENERATED_PURCHASE_REGISTER.iloc[i]['GSTIN/UIN']:
            statement = statement +' '+ 'GSTIN,'
        if ORIGINAL_GSTR2A.iloc[i]['Invoice Date'] != GENERATED_PURCHASE_REGISTER.iloc[i]['Date']:
            statement = statement +' '+ 'Invoice Date,'
        if ORIGINAL_GSTR2A.iloc[i]['Invoice Value(₹)'] != GENERATED_PURCHASE_REGISTER.iloc[i]['Gross Total']:
            statement = statement +' '+ 'Invoice Value,'
            print(ORIGINAL_GSTR2A.iloc[i]['Invoice Value(₹)'],GENERATED_PURCHASE_REGISTER.iloc[i]['Gross Total'])
        if ORIGINAL_GSTR2A.iloc[i]['Central Tax(₹)'] != GENERATED_PURCHASE_REGISTER.iloc[i]['C-GST 9%']:
            statement = statement +' '+ 'C GST,'
        if ORIGINAL_GSTR2A.iloc[i]['State/UT Tax(₹)'] != GENERATED_PURCHASE_REGISTER.iloc[i]['S-GST 9%']:
            statement = statement +' '+ 'State/UT Tax,'
        if ORIGINAL_GSTR2A.iloc[i]['Integrated Tax(₹)'] != GENERATED_PURCHASE_REGISTER.iloc[i]['I-GST 18%']:
            statement = statement +' '+ 'I GST,'
            
        remarks.append(statement)
        statement = ''
        # ws[f'W{i+1}'] = statement
        # wb.save('052024_27ADFPS4504B1ZV_GSTR2B_14062024 (1) (2).xlsx')



    # index1 = 2
    # index2 = 0
    
    # while index1 < len(remarks) or index2 < len(remarks):
    #     print(index1, index2)
    #     ws[f'W{index1}'] = remarks[index2]
    #     index2+=1
    #     index1+=1
        


    # ws['W1'] = 'Remarks'
    # wb.save(file2.filename)


    wb = load_workbook('CLIENT.xlsx')
    ws = wb.active
    
   
    index1 = 2
    index2 = 0
    
    while index1 < len(remarks) or index2 < len(remarks):
        print(index1, index2)
        ws[f'I{index1}'] = remarks[index2]
        index2+=1
        index1+=1
        

    ws['I1'] = 'Remarks'
    ws.insert_cols(1)
    wb.save('CLIENT.xlsx')


    df1 = pd.read_excel(file2.filename , sheet_name='B2B')
    df2 = pd.read_excel('CLIENT.xlsx')
    merged_df = pd.concat([df1, df2], ignore_index=False , axis=1)

    merged_df.to_excel('GSTR_ITR_RECO.xlsx', index=False, sheet_name='MergedSheet')





    print(remarks[0])

    df = pd.DataFrame(remarks)

    df.to_excel('remarks.xlsx')

    return True
